# -*- coding = utf-8 -*-
# @Time : 2021/5/287:59
# @Author : Sun Guijue
# @File : EHS_Exchange.py
# @Software : PyCharm

import datetime
import time
import serial
import modbus_tk.defines as cst
from modbus_tk import modbus_rtu
import pymssql
from openpyxl import load_workbook
import os

hostname = '172.30.100.109'
dbname = 'test_pdm'
username = 'zhangxiaoyi'
pwdstring = 'Zxy0903@'
sheetname = 'pdm_tempdata'

def Modbus_Read(slavenum:int,staradd:int,redtext:list,PORT:str):
    alarmtext = ""
    try:
        # 定义MODBUS通讯口，这里使用COM2端口（以为是模拟端口，COM2端口接收来自COM3的数据
        
        master = modbus_rtu.RtuMaster(serial.Serial(port=PORT, baudrate=9600, bytesize=8, parity='N', stopbits=1))
        master.set_timeout(10.0)
        master.set_verbose(True)
        # 上位机为主站，主站读取从站slave=slavenum的站点的信息，读取保持寄存器位
        redtext = master.execute(slave=slavenum, function_code=cst.READ_INPUT_REGISTERS, starting_address=staradd, quantity_of_x=1)
        alarmtext = "Connection OK"
        master.close()
        # return redtext, alarmtext
    except Exception as exc:
        print(str(exc))
        alarmtext = str(exc)
    return redtext, alarmtext

def SQL_Write(sheetname1:str,f_num:float,eq_no:str,set_num:str):  
    global username,hostname,pwdstring,dbname
    datetxt = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = pymssql.connect(server=hostname, user=username, password=pwdstring, database=dbname)
    try:
        cur = conn.cursor()
    except:
        print('Can connect to the database')
        return
    sqlstring = "update " + sheetname1 + " set time ='"+ datetxt +"'," + set_num + "='"+ str(format(f_num,'.3f')) +"' where eq_Codes ='"+ eq_no +"'"
    cur.execute(sqlstring)
    conn.commit()     # 必不可少！！
    cur.close()
    conn.close()

def Temp_Trans1(tempnum1:int):
    temp_t_num1 = tempnum1 * 0.079365 - 125
    if temp_t_num1 < 0:
       return 0
    else:
       return temp_t_num1

def Temp_Trans2(tempnum2:int):
    temp_t_num2 = tempnum2 * 0.048840 - 80
    if temp_t_num2 < 0:
        return 0
    else:
        return temp_t_num2

def Temp_Trans3(tempnum3:int):
    temp_t_num3 = tempnum3 * 0.061050 - 50
    if temp_t_num3 < 0:
        return 0
    else:
        return temp_t_num3

def Vib_Trans(vibnum:int):
    vib_t_num = vibnum * 0.030525 + 25
    return vib_t_num

def Cur_Trans1(curnum1:int):
    cur_t_num1 = curnum1 * 0.030525 - 25
    if cur_t_num1 < 0:
        return 0
    else:
        return cur_t_num1

def Cur_Trans2(curnum2:int):
    cur_t_num2 = curnum2 * 0.004762 - 3.9
    if cur_t_num2 < 0:
        return 0
    else:
        return cur_t_num2

def Sound_Trans(sudnum:int):
    sud_t_num = sudnum * 0.02747 + 7.5
    return sud_t_num

def Oil_Trans(oilnum:int):
    oil_t_num = oilnum * 0.006105 - 5
    if oil_t_num < 0:
        return 0
    else:
        return oil_t_num

from goto import with_goto
@with_goto
def main(timegap:int):
    global sheetname, rdtxt
    while True:
        ## 读取配置文件
        wb = load_workbook("ConfigFile.xlsx")
        ws = wb["AI_PDM"]
        filenum = ws.max_row
        now_h = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(now_h)
        for num in range (2,filenum + 1):
            
            rdtext = []
            PORT = ws.cell(row = num, column = 1).value
            SLAVE = ws.cell(row = num, column = 2).value
            ADD = ws.cell(row = num, column = 3).value - 1
            EQNO = ws.cell(row = num, column = 4).value
            TYPE = ws.cell(row = num, column = 5).value
            SET1 = ws.cell(row = num, column = 6).value
            
            label.begin
            try:
                rdtxt = Modbus_Read(int(SLAVE),int(ADD),rdtext,PORT)[0][0]
                print(rdtxt)
            except:
                   print("通讯中断！")
                   goto.begin
                
            if TYPE == 'T1':
                rdnum = Temp_Trans1(int(rdtxt))
            if TYPE == 'T2':
                rdnum = Temp_Trans2(int(rdtxt))
            if TYPE == 'T3':
                rdnum = Temp_Trans3(int(rdtxt))
            if TYPE == 'C1':
                rdnum = Cur_Trans1(int(rdtxt))
            if TYPE == 'C2':
                rdnum = Cur_Trans2(int(rdtxt))
            if TYPE == 'V1':
                rdnum = Vib_Trans(int(rdtxt))
            if TYPE == 'S1':
                rdnum = Sound_Trans(int(rdtxt))
            if TYPE == 'O1':
                rdnum = Oil_Trans(int(rdtxt))
        
            SQL_Write(sheetname,rdnum, str(EQNO), str(SET1))
            time.sleep(2)
           
            
        print("Already Update all the data in SQL DataBase")
        time.sleep(timegap)

if __name__ == '__main__':
    timegap = 60 * 5         # update the database every 5 mins
    print("Program Running ...")
    main(timegap)
